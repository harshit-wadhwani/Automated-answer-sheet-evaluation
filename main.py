from flask import Flask, render_template, redirect, send_file, url_for,request,jsonify, session, send_from_directory
from utilities.dbmanager import dbmanager
from utilities.qpapergenerator import GenerateQpaper
from utilities.ocrmanager import detect_document_text
from utilities.tempmanager import checkscore
from utilities.scoremanager import check_similarity
import json
from werkzeug.utils import secure_filename
import os
import hashlib
from huggingface_hub import from_pretrained_keras
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment


app = Flask(__name__)

app.config['UPLOAD_FOLDER'] = 'uploads/'
model = from_pretrained_keras("keras-io/bert-semantic-similarity")

# Compile the model

app.secret_key = 'your_secret_key'

@app.route("/results", methods=["GET", "POST"])
def results():
    return render_template("results.html")

@app.route("/", methods=["GET","POST"])
def home():
    return render_template("new_index.html")

@app.route("/createquestions", methods=['GET', 'POST'])
def createquestions():
    return render_template("index.html")

@app.route('/qp.html')  # Define route for qp.html
def qp():
    
    return render_template('qp.html')

@app.route('/data', methods=['POST'])
def save_data():
    
    data = request.get_json()
    with open('data/current.json', 'w') as f:
        json.dump(data, f)
    return jsonify({'message': 'Data saved successfully'})

@app.route('/generate', methods=['POST'])
def generate():
    db_manager = dbmanager()
    qp_generater= GenerateQpaper()
    collection_name='questions'
    with open('data/current.json', 'r') as file:
        json_data = json.load(file)
    json_data = json_data['data']
    new_json_data = {}
    for item in json_data:
        date= item['date']
        code = item['code']
        field = code+date
        if field not in new_json_data:
            new_json_data[field] = []
            
        del item['code']
        del item['date']
        new_json_data[field].append(item)

    db_manager.create(collection_name,new_json_data)
    qp_generater.MakePaper()

    message = "Python function executed successfully!"
    return jsonify({"message": message})


@app.route('/download.html')  # Define route for qp.html
def download():
    return render_template('download.html')

@app.route('/generate', methods=['GET'])
def get_generate_result():
    return send_file('data/questionpaper.pdf', as_attachment=True)

@app.route('/download/this.pdf')
def download_this_pdf():
    return send_file('data/questionpaper.pdf', as_attachment=True)


    
@app.route("/evaluation", methods=['GET', 'POST'])
def evaluation():
    return render_template("evaluation.html")

@app.route("/previous", methods=['GET', 'POST'])
def previous():
    excel_files = os.listdir('outputs')
    return render_template("previous.html", excel_files=excel_files)

@app.route('/get/<filename>')
def get_file(filename):
    return send_from_directory('outputs', filename, as_attachment=True)


@app.route('/upload', methods=['POST', 'GET'])
def upload_file():
    session["code"] = request.form.get("sub_code")
    temp_date = request.form.get("exam_date")
    # date_obj = datetime.strptime(temp_date, '%Y-%m-%d')
    # session["dateup"] = date_obj.strftime('%d-%m-%Y')
    session["dateup"] = str(temp_date)
    print(session["code"])
    print(session["dateup"])
    file_names = []
    for f in request.files.getlist('file'):
        if f.filename != '':
            filename = secure_filename(f.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            f.save(file_path)
            file_names.append(file_path.replace("uploads/", ""))
            
    result_id = "||".join(file_names)
    return redirect(url_for("result", result_id=result_id))

@app.route('/download_result')
def download_file():
    # Path to the file you want to download
    file_path = f"outputs\\results-{session['code']}{session['dateup']}.xlsx"
    # Specify the filename for the downloaded file
    filename = f"RESULTS-{session['code']}{session['dateup']}.xlsx"
    # Send the file to the client for download
    return send_file(file_path, as_attachment=True )

@app.route("/download_scores")
def download_scores():
    file_path = f"outputs\\scores-{session['code']}{session['dateup']}.xlsx"
    
    return send_file(file_path, as_attachment=True)


@app.route("/result/<result_id>")
def result(result_id):
    
    file_nms =  result_id.split("||")
    
    l_info_list = []
    l_ans_list = []
    l_pat_list = []
    # l_scr_list = []
    scores= []
    db_client = dbmanager()
    
    for filename in file_nms:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        # print(file_path)
        
        l_info,l_ans,l_pat= detect_document_text(file_path)
        scr_temp = []
        query = str(session["code"] + session["dateup"])
        print(query)
        
        ques, ref_ans, scores_assigned = db_client.get_quenum_ans_dict("questions", query)
            
        
        for ans, ref_answer, scores_q in zip(l_ans, ref_ans, scores_assigned):
            que_scr = check_similarity(ans, ref_answer, model)
            scr_temp.append(round(abs(que_scr['Perfect'] - que_scr['Contradiction']),2)*float(scores_q))
        
        scores.append(scr_temp)
        
            
        data = {}
        
        for i in range(len(l_pat)):

            uni_key = l_info[2]+"-"+l_info[3]

            if uni_key not in data:
                data[uni_key] = {}

            if l_info[1] not in data[uni_key]:
                data[uni_key][l_info[1]] = []

            data[uni_key][l_info[1]].append({
                "quenum": l_pat[i],
                "ans":  l_ans[i] ,
                "score" : scr_temp[i]
            })

        db_manager = dbmanager()
        collection_name='answers'
        db_manager.create(collection_name,data)
        l_info_list.append(l_info)
        l_ans_list.append(l_ans)
        l_pat_list.append(l_pat)
        
    usn_list = [i[2] for i in l_info_list]
    name_list = [i[0] for i in l_info_list]
    
    no_ques = len(l_pat_list[0])
    
    d = {}
    
    d["USN"] = usn_list
    d["Name"] = name_list
    
    for i in range(no_ques):
        d[f"q{i+1}"] = [a[i] for a in l_ans_list]
        
        
    s = {}
    s["USN"] = usn_list
    s["Name"] = name_list
    
    for i in range(no_ques):
        s[f"q{i+1}"] = [a[i] for a in scores]
    
    total_score = [round(sum(score)) for score in scores]
    s["Student's Score"] = total_score
    
    # s["Total Score"] = [ sum(scores_assigned) for i in range(no_ques)]
    
    df_scores = pd.DataFrame(s)
    df = pd.DataFrame(d)
    df_scores.to_excel(f"outputs/scores-{query}.xlsx")
    df.to_excel(f"outputs/results-{query}.xlsx")
    adjust_excel_formatting(f"outputs/scores-{query}.xlsx")
    adjust_excel_formatting(f"outputs/results-{query}.xlsx")
    
    return render_template('result.html', l_info=l_info_list, l_ans=l_ans_list, score =scores)



def adjust_excel_formatting(file_path, min_row_height=12, divisor=5):
    try:
        # Load the Excel file
        wb = load_workbook(file_path)
        ws = wb.active  # Get the active worksheet

        # Iterate over each row in the worksheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                # Set text wrap
                cell.alignment = Alignment(wrap_text=True)
                # Adjust row height based on content length
                if cell.value:
                    content_length = len(str(cell.value))
                    # Calculate the row height using a minimum height or divisor
                    new_height = max(content_length / divisor, min_row_height)
                    cell_row = ws.row_dimensions[cell.row]
                    cell_row.height = new_height

        # Save the changes to the Excel file
        wb.save(file_path)
        print("Formatting adjusted successfully.")
    except Exception as e:
        print(f"Error occurred: {str(e)}")


if __name__ == '__main__':
    app.run(debug=True)